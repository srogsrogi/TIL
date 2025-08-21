# openpyxl : excel을 제어하는 파이썬 패키지
# pystudy_env 가상환경에서 openpyxl 설치. conda install openpyxl

import openpyxl as op
print(op.__version__)

# 1. workbook : 엑셀파일 하나(.xls .xlsx 등)와 대응하는 객체(instance)
# 2. worksheet : workbook 하위의 엑셀파일 시트와 대응하는 객체. 이름 또는 인덱스로 조회 가능
# 3. cell : 엑셀방식(열-행)의 조회(ex. b3) 가능, 인덱스방식 조회(ex. (4,3)) 가능
# 4. range : 연속된 셀의 묶음. 행-열 순서대로 접근 가능

# workbook
# 새 workbook개체를 생성할 수 있다
# 기존 엑셀파일을 불러와서 workbook개체를 만들 수 있다

# 실제 액셀앱을 통해 수식을 저장해야 한다(openpyxl의 제약). 파이참 내에서 실제 값을 보고 싶으면 엑셀 깔고 이거 해야됨
# import win32com.client
#
# excel = win32com.client.Dispatch('Excel.Application')
# abs_path = 'C:/Workspaces/rpa_workspace/playdata2/_01_rpa/_01_openpyxl/지출.xlsx'
# excel.Workbooks.Open(abs_path).Save()
# excel.quit()

def create_workbook():
    wb = op.Workbook()
    print(type(wb), wb)
    # 메모리상의 작업 후에 반드시 파일로 지정해야 한다
    wb.save('hello_openpyxl.xlsx')

def load_workbook(path):
    wb = op.load_workbook(path)
    print(wb)

 # 상대경로(현재 파일 또는 프로젝트 루트 디렉토리 기준). ./는 현재경로라는 뜻, 지금은 생략해도 똑같이 작동
path = './hello_openpyxl.xlsx'
load_workbook(path)

# 절대경로. 윈도우의 경로 구분자 \는 escaping하기 위해 \\로 바꾸거나, /로 바꿔도 알아들음
path = 'C:\\Workspaces\\rpa_workspace\\KangHangyeol\\_01_rpa\\_01_openpyxl\\hello_openpyxl.xlsx'

# 소프트코딩. 나중에 경로가 바뀌어도 알아서 작동
import os
# current working directory
# cwd = os.getcwd()
# print(cwd)
# path = cwd + '\\hello_openpyxl,xlsx'

# worksheet
# 새 시트를 생성, 기존 시트 불러오기, 삭제 가능
def get_worksheet(path):
    wb = op.load_workbook(path)

    # 새 시트 생성
    ws = wb.create_sheet('hello')
    ws = wb.create_sheet('world')

    print(wb.worksheets)

    wb.save(path) # path 넣으면 이 파일 경로 들어가는거

# get_worksheet(path)

print('-'*30)

def get_worksheet2(path):
    wb = op.load_workbook(path)
    # 현재 활성화된 시트
    ws = wb.active
    # 다른 시트 가져오기
    ws = wb['hello']
    print(ws)
# get_worksheet2(path)

# cell 가져오기
def get_cell(path):
    wb = op.load_workbook(path)
    ws = wb.active # 저장시 마지막에 활성화돼있었던 시트(커서가 있는 곳)을 불러옴
    # 방법1 : excel 방식
    cell = ws['B1']
    print(cell)
    print(cell.value) # 셀에 할당된 값 확인
    cell.value = 'coffee'
    wb.save(path)
    print(cell)
    # 방법 2: 행렬 인덱스
    cell = ws.cell(2,3)
    print(cell)
    cell.value = 'bread'
    wb.save(path)
    print(cell.value)
get_cell(path)

# range : 연속된 cell에 대한 그룹객체

def get_range(path):
    wb = op.load_workbook(path)
    ws = wb.active

    # range. 변수이름을 range로 하면 built-in함수 range를 못 쓴다
    rng = ws['A1:C2'] # 각 행의 셀들을 하나의 튜플로 묶고, 다른 행끼리는 다른 튜플로 묶임.
    # 즉 바깥 튜플의 요소의 개수가 열의 개수일거고, 안에 있는 튜플들의 요소의 개수가 행의 개수일 것
    print(rng)

    # 반복문으로 모든 cell을 순회하려면 이런 식으로 중첩반복문을 사용해야 한다
    for row in rng:
        print(row)
        for cell in row:
            print(cell)

    # range 안에 있는 셀에 등차수열 입력하기
    n = 1
    for row in rng:
        for cell in row:
            cell.value = n
            n += 1
    wb.save(path)
    # range 안에 있는 모든 셀에 할당된 값 출력
    for row in rng:
        print(row)
        for cell in row:
            print(cell)

# 'hello' 시트에 A1 ~ A100까지 100~1의 숫자를 차례로 작성. 잘안되네
# def set_numbers(path):
#     wb = op.load_workbook(path)
#     ws = wb['hello']
#     n = 100
#     for cell in range(101):
#         ws[-n,1]
#         n -= 1
#     wb.save(path)
#     for cell in range(101):
#         print(cell.value)

    # range로하기
def set_numbers(path):
    wb = op.load_workbook(path)
    ws = wb.worksheets[1] # wb.sheetnames[1], ws = 'hello' 다 됨
    rng = ws['A1:A100']
    n = 0
    for row in rng:
        cell = row[0]
        cell.value = 100 - n
        n += 1
    wb.save(path)

    # cell로하기
    for i in range(1,101):
        ws.cell(i,1).value = n
        n -= 1
    wb.save(path)


# set_numbers(path)



# get_range(path)

# row, column

def test_row_column(path):
    wb = op.load_workbook(path)
    ws = wb.active
    print(ws)
    print(ws.rows) # 이러면 generator가 나오는데, 일단 이건 반복문 돌리면 된다고 생각하면 됨

# rows 데이터가 있는 모든 행을 range객체처럼 반환. 데이터가 없는 빈 셀도 포함
    for row in ws.rows:
        for cell in row:
            print(cell.value, end = ' ')
        print()

    # column
    for col in ws.columns:
        # print(col)
        for cell in col:
            print(cell.value, end = ' ')




path = r'고기.xlsx' # 경로에 한글 들어가있으면 r 추가해주면 되긴 하는데 웬만하면 한글 쓰지 마
test_row_column(path)

# 엑셀 빌트인 함수
def function_sum1(path):
    wb = op.load_workbook(path)
    ws = wb.active
    print(ws)
    print(ws['E2'].value)
    ws['E2'].value = '=sum(C:C)' # C column의 모든 셀의 합
    wb.save(path)

    print(ws['E2'].value)
    wb.save(path)

function_sum1(path)


path = r'지출.xlsx'
function_sum1(path)

print('-'*30)

def function_product(path):
    wb = op.load_workbook(path)
    ws = wb.active

    # 최대행, 최대열
    print(ws.max_row)
    print(ws.max_column)

    for n in range(2, ws.max_row+1):
        ws[f'E{n}'].value = f'=C{n}*D{n}'
        print(ws[f'E{n}'].value)
    wb.save(path)
    print(ws[f'E{n}'].value)

path = r'간식비.xlsx'
function_product(path)